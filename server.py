from __future__ import annotations

import csv
import io
import json
import mimetypes
import os
import re
import sys
import tempfile
import traceback
import urllib.parse
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - startup guard
    raise SystemExit("openpyxl is required to import Excel files. Use the bundled Codex Python runtime.") from exc


ROOT = Path(__file__).resolve().parent
WEB_ROOT = ROOT / "web"
DATA_ROOT = ROOT / "data"
AUDIT_STATE_PATH = DATA_ROOT / "audit_state.json"
DATASET_PATH = DATA_ROOT / "current_dataset.json"

BOARD_STATUSES = [
    {"id": "pending_review", "label": "Pendiente de Revision"},
    {"id": "in_review", "label": "En Revision"},
    {"id": "reviewed_ok", "label": "Revisado - OK"},
    {"id": "needs_reconciliation", "label": "Necesita Conciliacion"},
    {"id": "pending_system_changes", "label": "Pendiente de Cambios en Sistema"},
    {"id": "blocked", "label": "Revision Presidencia"},
    {"id": "unassigned_credits", "label": "Creditos / Recibos sin Asignar"},
    {"id": "incobrable_legal", "label": "Incobrable / Legal"},
    {"id": "completed", "label": "Completado"},
]

ISSUE_TYPES = [
    "None",
    "Duplicated receipt",
    "Unapplied receipt",
    "Missing movement",
    "Credit note mismatch",
    "Debit note mismatch",
    "Customer dispute",
    "System timing difference",
    "Write-off review",
    "Other",
]

DEFAULT_AUDIT = {
    "status": "pending_review",
    "issueType": "None",
    "priority": "Normal",
    "owner": "",
    "followUpDate": "",
    "findingNote": "",
    "resolution": "",
    "expectedAdjustment": "",
    "reviewedBy": "",
    "reviewedAt": "",
    "tags": [],
    "auditLog": [],
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def read_json(path: Path, fallback):
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        backup = path.with_suffix(path.suffix + f".corrupt-{datetime.now().strftime('%Y%m%d%H%M%S')}")
        path.replace(backup)
        return fallback


def write_json(path: Path, payload) -> None:
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp.replace(path)


def excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date system used by openpyxl for raw numeric dates.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    return str(value)


def number(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_header(value) -> str:
    return text(value)


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def audit_state() -> dict:
    return read_json(AUDIT_STATE_PATH, {"clients": {}, "imports": []})


def dataset() -> dict:
    return read_json(
        DATASET_PATH,
        {
            "importedAt": "",
            "fileName": "",
            "sheetName": "",
            "headers": [],
            "clients": [],
            "summary": empty_summary(),
        },
    )


def empty_summary() -> dict:
    return {
        "clients": 0,
        "documents": 0,
        "totalPending": 0,
        "negativeClients": 0,
        "mixedClients": 0,
        "unassignedClients": 0,
        "over360Clients": 0,
    }


def initial_status_for(client: dict) -> str:
    if client["flags"]["missingClient"]:
        return "unassigned_credits"
    if client["flags"]["negativeOnly"] or client["flags"]["mixedBalance"]:
        return "needs_reconciliation"
    return "pending_review"


def document_year(document: dict) -> int | None:
    value = document.get("year")
    if isinstance(value, (int, float)):
        return int(value)
    date_value = document.get("date")
    if isinstance(date_value, str) and len(date_value) >= 4 and date_value[:4].isdigit():
        return int(date_value[:4])
    return None


def enrich_activity_flags(client: dict) -> bool:
    flags = client.setdefault("flags", {})
    documents = client.get("documents", [])
    has_invoice_2019_forward = any(
        text(document.get("type")).lower() == "factura" and (document_year(document) or 0) >= 2019
        for document in documents
    )
    has_pre_2019_balance = any((document_year(document) or 9999) < 2019 for document in documents)
    updates = {
        "hasInvoice2019Forward": has_invoice_2019_forward,
        "oldCarryforwardWithNewActivity": has_invoice_2019_forward and has_pre_2019_balance,
        "oldOnlyBalance": has_pre_2019_balance and not has_invoice_2019_forward,
    }
    changed = False
    for key, value in updates.items():
        if flags.get(key) != value:
            flags[key] = value
            changed = True
    return changed


def refresh_summary(payload: dict) -> bool:
    clients = payload.get("clients", [])
    old_summary = payload.get("summary", {})
    new_summary = {
        "clients": len(clients),
        "documents": sum(int(client.get("documentCount") or 0) for client in clients),
        "totalPending": round(sum(number(client.get("totalPending")) for client in clients), 2),
        "negativeClients": sum(1 for item in clients if item.get("flags", {}).get("negativeOnly")),
        "mixedClients": sum(1 for item in clients if item.get("flags", {}).get("mixedBalance")),
        "unassignedClients": sum(1 for item in clients if item.get("flags", {}).get("missingClient")),
        "over360Clients": sum(1 for item in clients if item.get("flags", {}).get("over360")),
        "invoice2019ForwardClients": sum(1 for item in clients if item.get("flags", {}).get("hasInvoice2019Forward")),
        "oldOnlyClients": sum(1 for item in clients if item.get("flags", {}).get("oldOnlyBalance")),
    }
    if old_summary != new_summary:
        payload["summary"] = new_summary
        return True
    return False


def merge_audit(client: dict, state: dict) -> dict:
    saved = state.get("clients", {}).get(client["key"])
    if saved:
        merged = {**DEFAULT_AUDIT, **saved}
        client["audit"] = merged
        return client

    audit = {**DEFAULT_AUDIT, "status": initial_status_for(client)}
    audit["auditLog"] = [
        {
            "at": now_iso(),
            "type": "created",
            "message": f"Importado desde el reporte de AR y clasificado como {status_label(audit['status'])}.",
        }
    ]
    client["audit"] = audit
    state.setdefault("clients", {})[client["key"]] = audit
    return client


def status_label(status_id: str) -> str:
    return next((item["label"] for item in BOARD_STATUSES if item["id"] == status_id), status_id)


def parse_workbook(path: Path, filename: str) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_header(value) for value in next(rows)]
    index = {header: idx for idx, header in enumerate(headers)}

    required = ["Razon Social", "valor_pendiente", "valor_original", "documento", "Tipo Doc"]
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    groups: dict[str, dict] = {}
    documents = 0
    total_pending = 0.0

    for row_number, row in enumerate(rows, start=2):
        record = {header: row[index[header]] if index[header] < len(row) else None for header in headers}
        pending = number(record.get("valor_pendiente"))
        original = number(record.get("valor_original"))
        total_pending += pending
        documents += 1

        cod_empresa = text(record.get("Cod Empresa"))
        cliente_id = text(record.get("cliente_id"))
        razon_social = text(record.get("Razon Social"))
        empresa = text(record.get("Empresa"))
        key = cod_empresa or (f"{empresa}-{cliente_id}" if empresa and cliente_id else "") or f"unassigned-{slug(razon_social) or row_number}"

        if key not in groups:
            groups[key] = {
                "key": key,
                "codEmpresa": cod_empresa,
                "clientId": cliente_id,
                "name": razon_social or "#N/A",
                "empresa": empresa,
                "contact": text(record.get("Contacto")),
                "address": text(record.get("Direccion1")),
                "legalName": text(record.get("Nombre")),
                "identification": text(record.get("identificacion")),
                "clientStatus": text(record.get("STATUS")),
                "totalPending": 0.0,
                "totalOriginal": 0.0,
                "documentCount": 0,
                "positiveDocs": 0,
                "negativeDocs": 0,
                "creditDocs": 0,
                "invoiceDocs": 0,
                "oldestDate": None,
                "newestDate": None,
                "maxDaysOverdue": None,
                "largestPendingDocument": 0.0,
                "agingBuckets": {},
                "docTypes": {},
                "invoiceStatuses": {},
                "documents": [],
                "flags": {},
            }

        client = groups[key]
        fecha = excel_date(record.get("fecha"))
        fecha_vencimiento = excel_date(record.get("fecha_vencimiento"))
        dias_vencidos = record.get("dias_vencidos")
        if isinstance(dias_vencidos, (int, float)):
            client["maxDaysOverdue"] = max(client["maxDaysOverdue"] or dias_vencidos, dias_vencidos)

        if fecha:
            client["oldestDate"] = min(client["oldestDate"] or fecha, fecha)
            client["newestDate"] = max(client["newestDate"] or fecha, fecha)

        tipo_doc = text(record.get("Tipo Doc")) or "Unknown"
        aging = text(record.get("Rango_Dias_Vencidos")) or "Unknown"
        status_factura = text(record.get("Status Factura")) or "Unknown"

        client["totalPending"] += pending
        client["totalOriginal"] += original
        client["documentCount"] += 1
        client["positiveDocs"] += 1 if pending > 0 else 0
        client["negativeDocs"] += 1 if pending < 0 else 0
        client["creditDocs"] += 1 if text(record.get("Origen")) == "C" or pending < 0 else 0
        client["invoiceDocs"] += 1 if tipo_doc.lower() == "factura" else 0
        if abs(pending) > abs(client["largestPendingDocument"]):
            client["largestPendingDocument"] = pending
        client["agingBuckets"][aging] = client["agingBuckets"].get(aging, 0) + 1
        client["docTypes"][tipo_doc] = client["docTypes"].get(tipo_doc, 0) + 1
        client["invoiceStatuses"][status_factura] = client["invoiceStatuses"].get(status_factura, 0) + 1
        client["documents"].append(
            {
                "row": row_number,
                "empresa": empresa,
                "year": record.get("Year"),
                "month": text(record.get("Mes")),
                "date": fecha,
                "dueDate": fecha_vencimiento,
                "document": text(record.get("documento")),
                "reference": text(record.get("referencia")),
                "type": tipo_doc,
                "origin": text(record.get("Origen")),
                "originalAmount": original,
                "pendingAmount": pending,
                "daysOverdue": number(dias_vencidos, None) if dias_vencidos not in (None, "") else None,
                "agingBucket": aging,
                "invoiceStatus": status_factura,
            }
        )

    state = audit_state()
    clients = []
    for client in groups.values():
        client["totalPending"] = round(client["totalPending"], 2)
        client["totalOriginal"] = round(client["totalOriginal"], 2)
        client["largestPendingDocument"] = round(client["largestPendingDocument"], 2)
        client["flags"] = {
            "missingClient": not client["codEmpresa"] or client["name"] in ("", "#N/A"),
            "negativeOnly": client["negativeDocs"] > 0 and client["positiveDocs"] == 0,
            "mixedBalance": client["negativeDocs"] > 0 and client["positiveDocs"] > 0,
            "hasCredits": client["creditDocs"] > 0,
            "over360": any(bucket.startswith("13 -") for bucket in client["agingBuckets"]),
            "inactive": client["clientStatus"].upper() == "INACTIVO",
        }
        enrich_activity_flags(client)
        client["documents"].sort(key=lambda item: (item["date"] or "", abs(item["pendingAmount"])), reverse=True)
        clients.append(merge_audit(client, state))

    clients.sort(key=lambda item: abs(item["totalPending"]), reverse=True)
    imported_at = now_iso()
    summary = {
        "clients": len(clients),
        "documents": documents,
        "totalPending": round(total_pending, 2),
        "negativeClients": sum(1 for item in clients if item["flags"]["negativeOnly"]),
        "mixedClients": sum(1 for item in clients if item["flags"]["mixedBalance"]),
        "unassignedClients": sum(1 for item in clients if item["flags"]["missingClient"]),
        "over360Clients": sum(1 for item in clients if item["flags"]["over360"]),
        "invoice2019ForwardClients": sum(1 for item in clients if item["flags"]["hasInvoice2019Forward"]),
        "oldOnlyClients": sum(1 for item in clients if item["flags"]["oldOnlyBalance"]),
    }
    state.setdefault("imports", []).append(
        {
            "fileName": filename,
            "sheetName": sheet.title,
            "importedAt": imported_at,
            "clients": len(clients),
            "documents": documents,
            "totalPending": round(total_pending, 2),
        }
    )
    state["imports"] = state["imports"][-20:]
    write_json(AUDIT_STATE_PATH, state)

    return {
        "importedAt": imported_at,
        "fileName": filename,
        "sheetName": sheet.title,
        "headers": headers,
        "statuses": BOARD_STATUSES,
        "issueTypes": ISSUE_TYPES,
        "clients": clients,
        "summary": summary,
    }


def board_payload() -> dict:
    payload = dataset()
    payload["statuses"] = BOARD_STATUSES
    payload["issueTypes"] = ISSUE_TYPES
    state = audit_state()
    changed = False
    for client in payload.get("clients", []):
        before = client.get("audit", {})
        changed = enrich_activity_flags(client) or changed
        merge_audit(client, state)
        changed = changed or before != client.get("audit", {})
    changed = refresh_summary(payload) or changed
    if changed:
        write_json(AUDIT_STATE_PATH, state)
        write_json(DATASET_PATH, payload)
    return payload


def update_client(key: str, updates: dict) -> dict:
    state = audit_state()
    clients = state.setdefault("clients", {})
    current = {**DEFAULT_AUDIT, **clients.get(key, {})}
    old_status = current.get("status")
    allowed_fields = {
        "status",
        "issueType",
        "priority",
        "owner",
        "followUpDate",
        "findingNote",
        "resolution",
        "expectedAdjustment",
        "reviewedBy",
        "reviewedAt",
        "tags",
    }
    for field, value in updates.items():
        if field in allowed_fields:
            current[field] = value
    current["updatedAt"] = now_iso()
    if current.get("status") != old_status:
        current.setdefault("auditLog", []).append(
            {
                "at": current["updatedAt"],
                "type": "status",
                "message": f"Movido de {status_label(old_status)} a {status_label(current.get('status'))}.",
            }
        )
    if updates.get("logMessage"):
        current.setdefault("auditLog", []).append(
            {"at": current["updatedAt"], "type": "note", "message": text(updates["logMessage"])}
        )
    clients[key] = current
    write_json(AUDIT_STATE_PATH, state)

    payload = dataset()
    for client in payload.get("clients", []):
        if client["key"] == key:
            client["audit"] = current
            break
    write_json(DATASET_PATH, payload)
    return current


def summary_rows_for_export(payload: dict) -> list[dict]:
    rows = []
    for client in payload.get("clients", []):
        audit = client.get("audit", {})
        rows.append(
            {
                "Cod Empresa": client.get("codEmpresa", ""),
                "Cliente ID": client.get("clientId", ""),
                "Razon Social": client.get("name", ""),
                "Empresa": client.get("empresa", ""),
                "Total Pendiente": client.get("totalPending", 0),
                "Documentos": client.get("documentCount", 0),
                "Fecha Mas Antigua": client.get("oldestDate", ""),
                "Dias Vencidos Max": client.get("maxDaysOverdue", ""),
                "Status Auditoria": status_label(audit.get("status", "")),
                "Tipo Hallazgo": audit.get("issueType", ""),
                "Prioridad": audit.get("priority", ""),
                "Responsable": audit.get("owner", ""),
                "Seguimiento": audit.get("followUpDate", ""),
                "Ajuste Esperado": audit.get("expectedAdjustment", ""),
                "Hallazgo": audit.get("findingNote", ""),
                "Resolucion": audit.get("resolution", ""),
                "Revisado Por": audit.get("reviewedBy", ""),
                "Revisado En": audit.get("reviewedAt", ""),
                "Flags": ", ".join([key for key, value in client.get("flags", {}).items() if value]),
            }
        )
    return rows


def document_rows_for_export(payload: dict) -> list[dict]:
    rows = []
    for client in payload.get("clients", []):
        audit = client.get("audit", {})
        for document in client.get("documents", []):
            rows.append(
                {
                    "Cod Empresa": client.get("codEmpresa", ""),
                    "Cliente ID": client.get("clientId", ""),
                    "Razon Social": client.get("name", ""),
                    "Empresa": client.get("empresa", ""),
                    "Status Auditoria": status_label(audit.get("status", "")),
                    "Tipo Hallazgo": audit.get("issueType", ""),
                    "Prioridad": audit.get("priority", ""),
                    "Hallazgo": audit.get("findingNote", ""),
                    "Resolucion": audit.get("resolution", ""),
                    "Documento": document.get("document", ""),
                    "Referencia": document.get("reference", ""),
                    "Tipo Doc": document.get("type", ""),
                    "Origen": document.get("origin", ""),
                    "Fecha": document.get("date", ""),
                    "Fecha Vencimiento": document.get("dueDate", ""),
                    "Valor Original": document.get("originalAmount", 0),
                    "Valor Pendiente": document.get("pendingAmount", 0),
                    "Dias Vencidos": document.get("daysOverdue", ""),
                    "Rango Dias Vencidos": document.get("agingBucket", ""),
                    "Status Factura": document.get("invoiceStatus", ""),
                }
            )
    return rows


def status_rows_for_export(payload: dict) -> list[dict]:
    rows = []
    clients = payload.get("clients", [])
    for status in BOARD_STATUSES:
        status_clients = [client for client in clients if client.get("audit", {}).get("status") == status["id"]]
        rows.append(
            {
                "Status": status["label"],
                "Clientes": len(status_clients),
                "Documentos": sum(int(client.get("documentCount") or 0) for client in status_clients),
                "Total Pendiente": round(sum(number(client.get("totalPending")) for client in status_clients), 2),
                "Clientes Negativos": sum(1 for client in status_clients if client.get("flags", {}).get("negativeOnly")),
                "Clientes Mixtos": sum(1 for client in status_clients if client.get("flags", {}).get("mixedBalance")),
                "Clientes 2019+": sum(1 for client in status_clients if client.get("flags", {}).get("hasInvoice2019Forward")),
            }
        )
    return rows


def write_rows_as_table(sheet, rows: list[dict], table_name: str) -> None:
    if not rows:
        sheet.append(["No data"])
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    max_row = sheet.max_row
    max_col = sheet.max_column
    ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
    table = openpyxl.worksheet.table.Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="14253D")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(vertical="center")

    amount_headers = {"Total Pendiente", "Valor Original", "Valor Pendiente", "Ajuste Esperado"}
    date_headers = {"Fecha", "Fecha Vencimiento", "Fecha Mas Antigua", "Seguimiento", "Revisado En"}
    for col_idx, header in enumerate(headers, start=1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        width = min(max(len(str(header)) + 4, 12), 42)
        for row_idx in range(2, min(max_row, 60) + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                width = min(max(width, len(str(value)) + 2), 42)
        sheet.column_dimensions[letter].width = width
        if header in amount_headers:
            for row_idx in range(2, max_row + 1):
                sheet.cell(row=row_idx, column=col_idx).number_format = '#,##0.00;[Red]-#,##0.00'
        if header in date_headers:
            for row_idx in range(2, max_row + 1):
                sheet.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"


def build_export_workbook(payload: dict) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.properties.title = "AR Audit Export"
    workbook.properties.subject = "Accounts receivable audit board export"
    workbook.properties.creator = "AR Audit Board"
    workbook.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)

    summary_sheet = workbook.active
    summary_sheet.title = "Audit Summary"
    write_rows_as_table(summary_sheet, summary_rows_for_export(payload), "AuditSummaryTable")

    documents_sheet = workbook.create_sheet("Documents")
    write_rows_as_table(documents_sheet, document_rows_for_export(payload), "DocumentsTable")

    status_sheet = workbook.create_sheet("Status Summary")
    write_rows_as_table(status_sheet, status_rows_for_export(payload), "StatusSummaryTable")

    meta_sheet = workbook.create_sheet("Export Info")
    meta_rows = [
        {"Field": "Source File", "Value": payload.get("fileName", "")},
        {"Field": "Source Sheet", "Value": payload.get("sheetName", "")},
        {"Field": "Source Imported At", "Value": payload.get("importedAt", "")},
        {"Field": "Exported At", "Value": now_iso()},
        {"Field": "Clients", "Value": payload.get("summary", {}).get("clients", 0)},
        {"Field": "Documents", "Value": payload.get("summary", {}).get("documents", 0)},
    ]
    write_rows_as_table(meta_sheet, meta_rows, "ExportInfoTable")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class Handler(BaseHTTPRequestHandler):
    server_version = "ARAuditBoard/1.0"

    def log_message(self, format, *args):  # noqa: A003
        sys.stdout.write("%s - %s\n" % (self.address_string(), format % args))

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_text(self, text_body: str, status=200, content_type="text/plain; charset=utf-8"):
        body = text_body.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):  # noqa: N802
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/board":
            return self.send_json(board_payload())
        if parsed.path == "/api/export.csv":
            payload = board_payload()
            output = io.StringIO()
            rows = summary_rows_for_export(payload)
            if rows:
                writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)
            body = output.getvalue().encode("utf-8-sig")
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="ar-audit-summary.csv"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path == "/api/export.xlsx":
            body = build_export_workbook(board_payload())
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="ar-audit-export.xlsx"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path == "/api/export.json":
            body = json.dumps(board_payload(), ensure_ascii=False, indent=2).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="ar-audit-board.json"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        return self.serve_static(parsed.path)

    def do_POST(self):  # noqa: N802
        parsed = urllib.parse.urlparse(self.path)
        try:
            if parsed.path == "/api/import":
                return self.handle_import()
            if parsed.path.startswith("/api/client/"):
                key = urllib.parse.unquote(parsed.path.removeprefix("/api/client/"))
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8")
                updates = json.loads(body or "{}")
                return self.send_json({"audit": update_client(key, updates)})
            return self.send_json({"error": "Not found"}, status=404)
        except Exception as exc:  # pragma: no cover - runtime diagnostics
            traceback.print_exc()
            return self.send_json({"error": str(exc)}, status=500)

    def serve_static(self, request_path: str):
        if request_path in ("", "/"):
            request_path = "/index.html"
        target = (WEB_ROOT / request_path.lstrip("/")).resolve()
        if WEB_ROOT.resolve() not in target.parents and target != WEB_ROOT.resolve():
            return self.send_text("Forbidden", status=403)
        if not target.exists() or not target.is_file():
            return self.send_text("Not found", status=404)
        content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        body = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def handle_import(self):
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            return self.send_json({"error": "Expected multipart/form-data upload."}, status=400)
        boundary_match = re.search("boundary=(.+)", content_type)
        if not boundary_match:
            return self.send_json({"error": "Missing upload boundary."}, status=400)
        boundary = boundary_match.group(1).strip('"')
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        file_name, file_bytes = extract_multipart_file(body, boundary)
        if not file_bytes:
            return self.send_json({"error": "No Excel file was uploaded."}, status=400)
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file_name).suffix or ".xlsx") as temp:
            temp.write(file_bytes)
            temp_path = Path(temp.name)
        try:
            parsed = parse_workbook(temp_path, file_name)
            write_json(DATASET_PATH, parsed)
            return self.send_json(parsed)
        finally:
            temp_path.unlink(missing_ok=True)


def extract_multipart_file(body: bytes, boundary: str) -> tuple[str, bytes]:
    marker = ("--" + boundary).encode("utf-8")
    for part in body.split(marker):
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        headers, _, payload = part.partition(b"\r\n\r\n")
        disposition = headers.decode("utf-8", errors="ignore")
        if "filename=" not in disposition:
            continue
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        filename = filename_match.group(1) if filename_match else "upload.xlsx"
        return filename, payload.rstrip(b"\r\n")
    return "upload.xlsx", b""


def main():
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"AR Audit Board running at http://127.0.0.1:{port}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()
